#!/usr/bin/env python3
"""
Подготовка данных из xlsx для классификации ТНВЭД.
Извлекает уникальные артикулы + фото → temp директория.

Использование:
  python3 prepare-batch.py input.xlsx /tmp/classifier-batch/

Выход:
  /tmp/classifier-batch/
    articles.jsonl    — по строке на артикул (текстовые данные)
    images/row_NNNN.png — фото привязанные к строкам
    image_map.json    — {row: filename}
"""

import json
import os
import sys
import zipfile
import xml.etree.ElementTree as ET

import openpyxl


def extract_images(xlsx_path: str, out_dir: str) -> dict:
    """Extract images from xlsx, return {row: filename}."""
    z = zipfile.ZipFile(xlsx_path)

    try:
        rels_xml = z.read('xl/drawings/_rels/drawing1.xml.rels')
    except KeyError:
        return {}

    rels_root = ET.fromstring(rels_xml)
    rid_to_file = {}
    for rel in rels_root:
        rid = rel.get('Id')
        target = rel.get('Target')
        if target and 'image' in target:
            rid_to_file[rid] = target.replace('../', 'xl/')

    drawing = z.read('xl/drawings/drawing1.xml')
    root = ET.fromstring(drawing)
    ns = {
        'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing',
        'a': 'http://schemas.openxmlformats.org/drawingml/2006/main',
        'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
    }

    img_dir = os.path.join(out_dir, 'images')
    os.makedirs(img_dir, exist_ok=True)

    image_map = {}
    anchors = root.findall('.//xdr:twoCellAnchor', ns) + root.findall('.//xdr:oneCellAnchor', ns)
    for anchor in anchors:
        fr = anchor.find('xdr:from', ns)
        if fr is not None:
            row = int(fr.find('xdr:row', ns).text)
            pic = anchor.find('.//xdr:pic', ns)
            if pic is not None:
                blip = pic.find('.//a:blip', ns)
                if blip is not None:
                    rid = blip.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed')
                    if rid in rid_to_file:
                        img_data = z.read(rid_to_file[rid])
                        ext = rid_to_file[rid].split('.')[-1]
                        fname = f'row_{row:04d}.{ext}'
                        with open(os.path.join(img_dir, fname), 'wb') as f:
                            f.write(img_data)
                        image_map[row] = fname

    return image_map


def extract_articles(xlsx_path: str, image_map: dict) -> list:
    """Extract unique articles with text data."""
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb[wb.sheetnames[0]]

    seen = set()
    articles = []

    for row_idx in range(2, ws.max_row + 1):
        article = ws.cell(row=row_idx, column=2).value
        if not article or article in seen:
            continue
        seen.add(article)

        # Find image for this article (check row and nearby)
        img_file = None
        for r in [row_idx - 1, row_idx, row_idx + 1]:
            if r in image_map:
                img_file = image_map[r]
                break

        articles.append({
            'row': row_idx,
            'article': str(article).strip(),
            'brand': ws.cell(row=row_idx, column=4).value or '',
            'name_ru': ws.cell(row=row_idx, column=5).value or '',
            'name_dt': ws.cell(row=row_idx, column=6).value or '',
            'name_fab': ws.cell(row=row_idx, column=38).value or '',
            'composition_ru': ws.cell(row=row_idx, column=9).value or '',
            'fabric_type': ws.cell(row=row_idx, column=13).value or '',
            'size': ws.cell(row=row_idx, column=15).value or '',
            'age': ws.cell(row=row_idx, column=17).value or '',
            'height': ws.cell(row=row_idx, column=20).value,
            'gender': ws.cell(row=row_idx, column=22).value or '',
            'country': ws.cell(row=row_idx, column=23).value or '',
            'tnved_existing': ws.cell(row=row_idx, column=25).value or '',
            'image_file': img_file,
        })

    return articles


def main():
    if len(sys.argv) < 3:
        print(f'Usage: {sys.argv[0]} input.xlsx output_dir')
        sys.exit(1)

    xlsx_path = sys.argv[1]
    out_dir = sys.argv[2]
    os.makedirs(out_dir, exist_ok=True)

    print(f'Extracting images from {xlsx_path}...')
    image_map = extract_images(xlsx_path, out_dir)
    print(f'  {len(image_map)} images extracted')

    print(f'Extracting articles...')
    articles = extract_articles(xlsx_path, image_map)
    print(f'  {len(articles)} unique articles')

    # Save as JSONL
    jsonl_path = os.path.join(out_dir, 'articles.jsonl')
    with open(jsonl_path, 'w', encoding='utf-8') as f:
        for a in articles:
            f.write(json.dumps(a, ensure_ascii=False) + '\n')

    # Save image map
    map_path = os.path.join(out_dir, 'image_map.json')
    with open(map_path, 'w') as f:
        json.dump({str(k): v for k, v in image_map.items()}, f, indent=2)

    print(f'Output: {jsonl_path}')
    print(f'Images: {out_dir}/images/')
    print(f'Map: {map_path}')


if __name__ == '__main__':
    main()
